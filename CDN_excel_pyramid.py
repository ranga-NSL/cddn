"""
Excel Pyramid Visualization Generator

Creates a visual pyramid structure in Excel Overview sheet showing the dependency network
hierarchy with fundamental and visible scores.
"""

import argparse
import os
from pathlib import Path
from typing import Dict, List, Any, Tuple
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Import from CDN_graph.py
from CDN_graph import DependencyNetwork, analyze_excel_structure, load_from_excel


def find_excel_files(directory: str = '.') -> List[str]:
    """Auto-detect all .xlsx files in directory."""
    path = Path(directory)
    excel_files = sorted([f.name for f in path.glob('*.xlsx') if not f.name.startswith('~$')])
    return excel_files


def select_excel_file_interactive() -> str:
    """Show list of Excel files and prompt user to select."""
    excel_files = find_excel_files()
    
    if not excel_files:
        raise FileNotFoundError("No .xlsx files found in current directory")
    
    print("\nAvailable Excel files:")
    for i, filename in enumerate(excel_files, 1):
        print(f"  {i}. {filename}")
    
    while True:
        try:
            choice = input(f"\nSelect file number (1-{len(excel_files)}): ").strip()
            index = int(choice) - 1
            if 0 <= index < len(excel_files):
                return excel_files[index]
            else:
                print(f"Please enter a number between 1 and {len(excel_files)}")
        except ValueError:
            print("Please enter a valid number")
        except KeyboardInterrupt:
            print("\nCancelled.")
            raise


def get_node_status(fundamental: float, visible: float) -> Tuple[str, str]:
    """
    Get status string and color for a node.
    Returns: (status_text, color_name)
    """
    if visible > fundamental:
        return f"INFLATED by {visible - fundamental:.2f}", "red"
    elif visible == fundamental:
        return "Aligned", "yellow"
    else:
        return f"Conservative by {fundamental - visible:.2f}", "green"


def abbreviate_description(description: str, max_length: int = 8) -> str:
    """Abbreviate description to short form like N-A, N-B, etc."""
    if not description or pd.isna(description):
        return ""
    
    desc_str = str(description).strip()
    
    # If already short, use as-is
    if len(desc_str) <= max_length:
        return desc_str
    
    # Try to extract meaningful abbreviation
    # If it contains "Node" or similar, extract letter/number
    words = desc_str.split()
    if len(words) > 1:
        # Try to get last word or meaningful part
        last_word = words[-1]
        if len(last_word) <= max_length:
            return last_word
    
    # Default: take first max_length characters
    return desc_str[:max_length]


def calculate_pyramid_layout(level_sizes: List[int], node_spacing: int = 4) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """
    Calculate cell positions for pyramid layout.
    Level 0 (base) at bottom, highest level at top.
    Returns: {(level, node_index): (row, col)}
    """
    positions = {}
    
    # Calculate total height needed (rows per level * number of levels)
    rows_per_level = 4  # 3 for content + 1 spacing
    total_levels = len(level_sizes)
    
    # Start from top - reserve space for title/legend
    start_row = 5  # Start after title and legend space
    
    # Calculate max width for centering
    max_width = max(level_sizes) * node_spacing if level_sizes else 10
    
    # Level 0 (base) should be at the BOTTOM (highest row number)
    # Highest level should be at the TOP (lowest row number)
    # So we iterate levels in reverse: highest level first (top), Level 0 last (bottom)
    
    for level in range(len(level_sizes)):
        level_size = level_sizes[level]
        
        # Calculate starting column to center this level
        level_width = level_size * node_spacing
        start_col = (max_width - level_width) // 2 + 1  # +1 for 1-based Excel columns
        
        # Calculate row position: 
        # Highest level (last in level_sizes) goes at top (start_row)
        # Level 0 (first in level_sizes) goes at bottom (start_row + (total_levels-1) * rows_per_level)
        # Reverse the order: level 0 gets highest row, highest level gets lowest row
        row_position = start_row + (total_levels - 1 - level) * rows_per_level
        
        # Position nodes in this level
        for node_idx in range(level_size):
            col = start_col + node_idx * node_spacing
            positions[(level, node_idx)] = (row_position, col)
    
    return positions


def create_pyramid_overview(excel_path: str):
    """Create/update Overview sheet with pyramid visualization."""
    print(f"\n=== Creating Excel Pyramid Visualization ===")
    print(f"Loading: {excel_path}")
    
    # Load network data
    net = load_from_excel(excel_path)
    net.compute()
    
    # Load workbook
    wb = load_workbook(excel_path)
    
    # Remove existing Overview sheet if it exists
    if 'Overview' in wb.sheetnames:
        wb.remove(wb['Overview'])
        print("  Removed existing Overview sheet")
    
    # Create new Overview sheet
    ws = wb.create_sheet('Overview', 0)  # Insert at beginning
    
    # Get descriptions from original sheets
    descriptions = {}
    excel_file = pd.ExcelFile(excel_path)
    
    for sheet_name in excel_file.sheet_names:
        if sheet_name.startswith('Level_'):
            try:
                level_num = int(sheet_name.split('_')[1])
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                desc_col = [col for col in df.columns if 'description' in col.lower() or 'desc' in col.lower()]
                if desc_col:
                    for idx, desc in enumerate(df[desc_col[0]]):
                        descriptions[(level_num, idx)] = abbreviate_description(desc)
            except:
                pass
    
    # Calculate layout
    positions = calculate_pyramid_layout(net.level_sizes)
    
    # Set default column width for ALL columns that might be used
    max_col = max([pos[1] for pos in positions.values()]) + 15  # Add buffer for legend
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        # Force set width - override any existing value
        ws.column_dimensions[col_letter].width = 20  # Wider default width
    
    # Define colors
    color_map = {
        'red': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),
        'yellow': PatternFill(start_color='FFFFE6', end_color='FFFFE6', fill_type='solid'),
        'green': PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid'),
    }
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write nodes
    for level in range(net.num_levels):
        level_size = net.level_sizes[level]
        
        for node_idx in range(level_size):
            row, col = positions[(level, node_idx)]
            
            # Get node data
            if level == 0:
                # Level 0: manual score only
                score = net.level0_manual[node_idx]
                desc = descriptions.get((level, node_idx), f"L{level}-{node_idx}")
                cell_text = f"{desc}\n{score:.2f}"
                fill_color = color_map['green']  # Base level is always good
            else:
                # Higher levels: fundamental and visible
                fund = net.level_fundamental[level-1][node_idx]
                vis = net.level_visible[level-1][node_idx]
                status_text, color_name = get_node_status(fund, vis)
                desc = descriptions.get((level, node_idx), f"L{level}-{node_idx}")
                cell_text = f"{desc}\nF:{fund:.2f}\nV:{vis:.2f}"
                fill_color = color_map[color_name]
            
            # Write to cell
            cell = ws.cell(row=row+1, column=col)
            cell.value = cell_text
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_style
            cell.fill = fill_color
            cell.font = Font(size=10, bold=True)
            
            # Set row height and column width (wider for better visibility)
            ws.row_dimensions[row+1].height = 60  # Increased from 45
            # Force column width - don't check if already set
            ws.column_dimensions[get_column_letter(col)].width = 20  # Increased to 20 for better visibility
    
    # Add level labels on the left
    for level in range(net.num_levels):
        row, _ = positions.get((level, 0), (level * 4, 1))
        label_cell = ws.cell(row=row+1, column=1)
        if level == 0:
            label_cell.value = f"LEVEL {level}\n(Base)"
        else:
            label_cell.value = f"LEVEL {level}"
        label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        label_cell.font = Font(size=11, bold=True)
        label_cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        label_cell.border = border_style
        ws.column_dimensions['A'].width = 15  # Increased from 12
        ws.row_dimensions[row+1].height = 60  # Ensure level label rows also have proper height
    
    # Add legend at the top
    legend_row = 1
    legend_start_col = max([pos[1] for pos in positions.values()]) + 3
    
    legend_title = ws.cell(row=legend_row, column=legend_start_col, value="Legend:")
    legend_title.font = Font(bold=True, size=11)
    
    legend_items = [
        ("Green", "Conservative (V < F)"),
        ("Yellow", "Aligned (V = F)"),
        ("Red", "Inflated (V > F)"),
        ("", "F = Fundamental score"),
        ("", "V = Visible score")
    ]
    
    for idx, (color, text) in enumerate(legend_items):
        row = legend_row + idx + 1
        if color:
            fill = color_map.get(color.lower(), None)
            legend_color_cell = ws.cell(row=row, column=legend_start_col, value=color)
            legend_color_cell.fill = fill
            legend_color_cell.border = border_style
            legend_color_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(legend_start_col)].width = 12
        legend_text_cell = ws.cell(row=row, column=legend_start_col + 1, value=text)
        legend_text_cell.font = Font(size=10)
        legend_text_cell.alignment = Alignment(vertical='center')
        ws.column_dimensions[get_column_letter(legend_start_col + 1)].width = 25
    
    # Add title
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Dependency Network Pyramid"
    title_cell.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 25  # Title row height
    
    # Save workbook
    wb.save(excel_path)
    print(f"  ✓ Overview sheet created successfully!")
    print(f"  ✓ Pyramid visualization saved to: {excel_path}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Generate Excel pyramid visualization for CDN dependency network'
    )
    parser.add_argument(
        'excel_file',
        nargs='?',
        help='Excel file path (optional - will prompt if not provided)'
    )
    args = parser.parse_args()
    
    try:
        if args.excel_file:
            excel_path = args.excel_file
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"File not found: {excel_path}")
        else:
            excel_path = select_excel_file_interactive()
        
        create_pyramid_overview(excel_path)
        print("\n✓ Complete!")
        
    except FileNotFoundError as e:
        print(f"\n✗ Error: {e}")
    except KeyboardInterrupt:
        print("\n\nCancelled by user.")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()

